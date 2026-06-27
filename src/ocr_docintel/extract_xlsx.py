# XLSX를 시트별 표로 추출하고 병합셀을 보존하는 파서 (PRD FR-3, System Prompt §3.3)
from __future__ import annotations
from pathlib import Path

import openpyxl

from .ir import Document, Element, Source


def _sheet_rows(ws):
    """시트를 2차원 문자열 그리드로. 병합셀은 대표(좌상단) 값만 두고 나머지는 빈 칸."""
    if ws.max_row == 0 or ws.max_column == 0:
        return []
    grid = [["" for _ in range(ws.max_column)] for _ in range(ws.max_row)]
    for r in range(ws.max_row):
        for c in range(ws.max_column):
            v = ws.cell(row=r + 1, column=c + 1).value
            grid[r][c] = "" if v is None else str(v).replace("\n", "<br>")

    # 병합 범위: 좌상단 외 셀을 빈 칸으로(FR-3 확장 구조 표현).
    for rng in ws.merged_cells.ranges:
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if r == rng.min_row and c == rng.min_col:
                    continue
                grid[r - 1][c - 1] = ""

    # 완전히 빈 후행/후열 트리밍.
    while grid and all(cell == "" for cell in grid[-1]):
        grid.pop()
    return grid


def extract(path: str | Path) -> Document:
    """XLSX → Document. 시트마다 `## Sheet: 이름` 제목 + 표 Element."""
    path = Path(path)
    wb = openpyxl.load_workbook(str(path), data_only=True)
    elements: list[Element] = []
    order = 0
    for ws in wb.worksheets:
        elements.append(Element(
            type="heading", order=order, level=2,
            content=f"Sheet: {ws.title}",
            source=Source(sheet=ws.title),
        ))
        order += 1
        rows = _sheet_rows(ws)
        if rows:
            elements.append(Element(
                type="table", order=order, level=0, content=rows,
                source=Source(sheet=ws.title),
            ))
            order += 1
    wb.close()
    return Document(document_title=path.stem, source_format="xlsx", elements=elements)
