# 웹 데모용 예시 문서를 reportlab/openpyxl로 즉석 생성하는 모듈
from __future__ import annotations
import io


def _pdf(draw) -> bytes:
    """reportlab canvas로 그린 PDF를 bytes로 반환."""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    draw(c, *A4)
    c.save()
    return buf.getvalue()


def _grid(data, col_widths, fontname="Helvetica"):
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    t = Table(data, colWidths=col_widths, rowHeights=22)
    t.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("FONTNAME", (0, 0), (-1, -1), fontname),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
    ]))
    return t


def _section_hierarchy() -> bytes:
    def draw(c, w, h):
        c.setFont("Helvetica-Bold", 22); c.drawString(72, h - 90, "Quarterly Operations Review")
        c.setFont("Helvetica-Bold", 15); c.drawString(72, h - 130, "1. Overview")
        c.setFont("Helvetica", 11); c.drawString(72, h - 156, "This section summarizes the quarter at a high level.")
        c.setFont("Helvetica-Bold", 13); c.drawString(72, h - 192, "1.1 Highlights")
        c.setFont("Helvetica", 11); c.drawString(72, h - 216, "Revenue grew while operating costs stayed flat.")
        c.setFont("Helvetica-Bold", 15); c.drawString(72, h - 252, "2. Outlook")
        c.setFont("Helvetica", 11); c.drawString(72, h - 278, "Guidance remains unchanged for the next period.")
    return _pdf(draw)


def _financial_table() -> bytes:
    def draw(c, w, h):
        c.setFont("Helvetica-Bold", 20); c.drawString(72, h - 90, "Form 10-K Financial Summary")
        data = [["Line Item", "FY2024", "FY2025"],
                ["Revenue", "1,204", "1,455"],
                ["Operating Income", "312", "398"],
                ["Net Income", "210", "265"]]
        t = _grid(data, [180, 90, 90]); t.wrapOn(c, w, h); t.drawOn(c, 72, h - 230)
    return _pdf(draw)


def _invoice() -> bytes:
    def draw(c, w, h):
        c.setFont("Helvetica-Bold", 22); c.drawString(72, h - 90, "INVOICE #2026-0042")
        c.setFont("Helvetica", 11); c.drawString(72, h - 120, "Bill To: Acme Corp.")
        data = [["Description", "Qty", "Amount"],
                ["Consulting", "10", "$2,000"],
                ["Support", "3", "$450"],
                ["Total", "", "$2,450"]]
        t = _grid(data, [200, 60, 90]); t.wrapOn(c, w, h); t.drawOn(c, 72, h - 230)
    return _pdf(draw)


def _science_table() -> bytes:
    def draw(c, w, h):
        c.setFont("Helvetica-Bold", 18); c.drawString(72, h - 90, "Reservoir Seepage Risk Measurements")
        data = [["Sensor", "Depth(m)", "Flow(L/s)"],
                ["S1", "2.4", "0.12"],
                ["S2", "5.1", "0.34"],
                ["S3", "8.0", "0.51"]]
        t = _grid(data, [120, 100, 100]); t.wrapOn(c, w, h); t.drawOn(c, 72, h - 220)
        c.setFont("Helvetica-Oblique", 10); c.drawString(72, h - 250, "Figure 1. Seepage flow by sensor depth.")
    return _pdf(draw)


def _xlsx(build) -> bytes:
    import openpyxl
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    build(wb)
    wb.save(buf)
    return buf.getvalue()


def _spreadsheet_empty() -> bytes:
    def build(wb):
        ws = wb.active; ws.title = "재고"
        ws["A1"] = "품목"; ws["B1"] = "수량"; ws["C1"] = "단가"
        ws["A2"] = "볼트"; ws["C2"] = 120          # B2 의도적 공란
        ws["A3"] = "너트"; ws["B3"] = 40            # C3 의도적 공란
    return _xlsx(build)


def _complex_table() -> bytes:
    def build(wb):
        ws = wb.active; ws.title = "손익"
        ws.merge_cells("A1:C1"); ws["A1"] = "2025 손익계산서"
        ws["A2"] = "구분"; ws["B2"] = "상반기"; ws["C2"] = "하반기"
        ws["A3"] = "매출"; ws["B3"] = 500; ws["C3"] = 620
        ws["A4"] = "비용"; ws["B4"] = 300; ws["C4"] = 340
    return _xlsx(build)


def _medical() -> bytes:
    def draw(c, w, h):
        c.setFont("Helvetica-Bold", 18); c.drawString(72, h - 90, "Patient Summary vs. Disclaimer")
        c.setFont("Helvetica-Bold", 13); c.drawString(72, h - 124, "Clinical Details")
        c.setFont("Helvetica", 11)
        c.drawString(72, h - 148, "Diagnosis: Type 2 diabetes, controlled with diet.")
        c.drawString(72, h - 168, "Medication: Metformin 500mg twice daily.")
        c.setFont("Helvetica-Bold", 13); c.drawString(72, h - 202, "Disclaimer")
        c.setFont("Helvetica-Oblique", 10)
        c.drawString(72, h - 224, "This summary is informational and not medical advice.")
    return _pdf(draw)


def _hearing() -> bytes:
    def draw(c, w, h):
        c.setFont("Helvetica-Bold", 18); c.drawString(72, h - 90, "City Council Hearing Testimony")
        c.setFont("Helvetica", 11)
        c.drawString(72, h - 124, "CHAIR: The hearing on housing policy will come to order.")
        c.drawString(72, h - 146, "WITNESS: Thank you for the opportunity to testify today.")
        c.drawString(72, h - 168, "WITNESS: Rents in the district rose 12 percent last year.")
    return _pdf(draw)


def _math_doc() -> bytes:
    def draw(c, w, h):
        c.setFont("Helvetica-Bold", 18); c.drawString(72, h - 90, "Quadratic Methods")
        c.setFont("Helvetica", 12)
        c.drawString(72, h - 124, "Given f(x) = a*x^2 + b*x + c, the roots are")
        c.drawString(72, h - 146, "x = (-b +/- sqrt(b^2 - 4ac)) / (2a).")
        c.drawString(72, h - 168, "The discriminant D = b^2 - 4ac decides the root count.")
    return _pdf(draw)


def _math_pdf() -> bytes:
    def draw(c, w, h):
        c.setFont("Helvetica-Bold", 18); c.drawString(72, h - 90, "Calculus Notes")
        c.setFont("Helvetica", 12)
        c.drawString(72, h - 124, "The derivative of x^n is n*x^(n-1).")
        c.drawString(72, h - 146, "Integral of x^n dx = x^(n+1)/(n+1) + C for n != -1.")
    return _pdf(draw)


def _unicode_pdf(title, lines, font_candidates, table=None) -> bytes:
    """Windows 유니코드 TTF를 등록해 비라틴 텍스트 PDF 생성. 폰트 부재 시 RuntimeError."""
    import os
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    fname = None
    for path, idx in font_candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("UNI", path, subfontIndex=idx)); fname = "UNI"; break
            except Exception:
                continue
    if not fname:
        raise RuntimeError("유니코드 폰트를 찾지 못했습니다(머신 의존)")

    def draw(c, w, h):
        c.setFont(fname, 18); c.drawString(72, h - 90, title)
        c.setFont(fname, 12)
        y = h - 130
        for ln in lines:
            c.drawString(72, y, ln); y -= 24
        if table:
            t = _grid(table, [140, 140], fontname=fname)
            t.wrapOn(c, w, h); t.drawOn(c, 72, y - 70)
    return _pdf(draw)


_JP_FONTS = [("C:/Windows/Fonts/meiryo.ttc", 0), ("C:/Windows/Fonts/msgothic.ttc", 0), ("C:/Windows/Fonts/YuGothR.ttc", 0)]
_HI_FONTS = [("C:/Windows/Fonts/Nirmala.ttf", 0), ("C:/Windows/Fonts/mangal.ttf", 0)]


def _japanese_table() -> bytes:
    return _unicode_pdf(
        "日本語のサンプル文書",
        ["これは日本語の本文です。", "下の表に売上をまとめます。"],
        _JP_FONTS,
        table=[["項目", "金額"], ["売上", "1200"], ["費用", "800"]],
    )


def _hindi() -> bytes:
    return _unicode_pdf(
        "हिंदी नमूना दस्तावेज़",
        ["यह एक हिंदी पाठ है।", "यह दूसरी पंक्ति है।"],
        _HI_FONTS,
    )


# id → (한글 파일명, content-type, builder). 이미지의 12종 예시와 대응.
SAMPLES = {
    "spreadsheet-empty": ("빈 셀 있는 스프레드시트.xlsx",
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", _spreadsheet_empty),
    "section-hierarchy": ("긴 문서의 섹션 계층 구조.pdf", "application/pdf", _section_hierarchy),
    "financial-10k":     ("금융 10K 테이블 추출.pdf", "application/pdf", _financial_table),
    "medical":           ("의료 세부사항 vs 면책 조항.pdf", "application/pdf", _medical),
    "hearing":           ("뉴욕시 시의회 청문회 증언.pdf", "application/pdf", _hearing),
    "complex-table":     ("복소 표.xlsx",
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", _complex_table),
    "math-doc":          ("수학 중심 문서.pdf", "application/pdf", _math_doc),
    "japanese-table":    ("일본어 + 표.pdf", "application/pdf", _japanese_table),
    "invoice":           ("송장.pdf", "application/pdf", _invoice),
    "hindi":             ("힌디어.pdf", "application/pdf", _hindi),
    "math-pdf":          ("수학 중심의 PDF.pdf", "application/pdf", _math_pdf),
    "science-pdf":       ("복잡한 표와 도표가 포함된 과학 PDF.pdf", "application/pdf", _science_table),
}


def build_sample(sample_id: str) -> tuple[str, str, bytes]:
    """샘플 id로 (파일명, content-type, bytes)를 생성. 미지원 id는 KeyError."""
    name, ctype, builder = SAMPLES[sample_id]
    return name, ctype, builder()
